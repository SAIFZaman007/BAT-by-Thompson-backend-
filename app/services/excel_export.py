"""Excel export builder for the admin dashboard.

Produces a single .xlsx workbook with three sheets:

  - "Submissions"      — one row per onboarding submission (core profile + status)
  - "KYC Documents"    — one row per uploaded document, linked back by reference code
  - "Support Messages" — one row per contact/support message, linked by email

This keeps every user's data in one downloadable file (per requirement #4)
while remaining easy to filter/sort in a spreadsheet tool.
"""
from __future__ import annotations

import datetime as dt
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.entities import ContactMessage, KycDocument, OnboardingSubmission

_HEADER_FILL = PatternFill(start_color="173EA5", end_color="173EA5", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _fmt_dt(value: dt.datetime | None) -> str:
    if value is None:
        return ""
    return value.strftime("%Y-%m-%d %H:%M:%S")


def _write_header(ws, headers: list[str]) -> None:
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws, headers: list[str], min_width: int = 12, max_width: int = 60) -> None:
    for col, title in enumerate(headers, start=1):
        longest = len(title)
        for row in ws.iter_rows(min_col=col, max_col=col, min_row=2):
            for cell in row:
                if cell.value is not None:
                    longest = max(longest, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = max(min_width, min(max_width, longest + 2))


def build_all_users_xlsx(
    submissions: list[OnboardingSubmission],
    kyc_documents: list[KycDocument],
    messages: list[ContactMessage],
) -> bytes:
    wb = Workbook()

    # ── Submissions ──────────────────────────────────────────────────────────
    ws_subs = wb.active
    ws_subs.title = "Submissions"
    sub_headers = [
        "Reference Code",
        "Full Name",
        "Email",
        "Phone",
        "Wallet Address",
        "Status",
        "Details",
        "Submitted At",
        "KYC Documents",
    ]
    _write_header(ws_subs, sub_headers)
    for sub in submissions:
        ws_subs.append(
            [
                sub.reference_code,
                sub.full_name,
                sub.email,
                sub.phone,
                sub.wallet_address,
                sub.status.value if hasattr(sub.status, "value") else str(sub.status),
                sub.details,
                _fmt_dt(sub.created_at),
                len(sub.kyc_documents),
            ]
        )
    _autosize(ws_subs, sub_headers)

    # ── KYC Documents ────────────────────────────────────────────────────────
    ws_kyc = wb.create_sheet("KYC Documents")
    kyc_headers = ["Reference Code", "User", "File Name", "MIME Type", "Uploaded At"]
    _write_header(ws_kyc, kyc_headers)
    sub_by_id = {sub.id: sub for sub in submissions}
    for doc in kyc_documents:
        owner = sub_by_id.get(doc.submission_id)
        ws_kyc.append(
            [
                owner.reference_code if owner else "",
                owner.full_name if owner else "",
                doc.original_name,
                doc.mime_type,
                _fmt_dt(doc.created_at),
            ]
        )
    _autosize(ws_kyc, kyc_headers)

    # ── Support Messages ─────────────────────────────────────────────────────
    ws_msgs = wb.create_sheet("Support Messages")
    msg_headers = ["Reference Code", "Name", "Email", "Message", "Sent At"]
    _write_header(ws_msgs, msg_headers)
    sub_by_email: dict[str, OnboardingSubmission] = {}
    for sub in submissions:
        sub_by_email.setdefault(sub.email.lower(), sub)
    for msg in messages:
        owner = sub_by_email.get(msg.email.lower())
        ws_msgs.append(
            [
                owner.reference_code if owner else "",
                msg.name,
                msg.email,
                msg.message,
                _fmt_dt(msg.created_at),
            ]
        )
    _autosize(ws_msgs, msg_headers)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
